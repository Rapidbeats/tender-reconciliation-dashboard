import pandas as pd
import numpy as np
import os
import sys
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Set
import warnings
import gc

warnings.filterwarnings('ignore')

# Try to import tkinter for file dialog
try:
    from tkinter import Tk, filedialog
    HAS_TKINTER = True
except ImportError:
    HAS_TKINTER = False

# Import openpyxl for formatting
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


class TenderReconciliationProcessor:
    def __init__(self, netting_threshold: float = 5.0, approval_filter: str = 'all'):
        """
        Initialize processor
        approval_filter: 'auto_approved_only' or 'all'
        """
        self.tender_mapping = {
            'cash': 'Cash',
            'card': 'Card',
            'upi': 'UPI',
            'wallet': 'Wallet'
        }

        self.column_mapping = {
            'store_id': ['store id', 'storeid', 'store_id'],
            'store_response': ['store response entry', 'store_response_entry', 'response entry'],
            'auto_approved': ['auto approved date', 'auto_approved_date', 'autoapproveddate']
        }

        self.netting_off_threshold = netting_threshold
        self.approval_filter = approval_filter

        self.classification_thresholds = [
            (0, 100, "Diff less than +/- 100"),
            (100, 1000, "Diff b/w +/- 1000"),
            (1000, 5000, "Diff b/w +/- 5000"),
            (5000, 10000, "Diff b/w +/- 10000"),
            (10000, 25000, "Diff b/w +/- 25000"),
            (25000, 50000, "Diff b/w +/- 50000"),
            (50000, float('inf'), "Diff more than +/- 50000")
        ]

        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(name='Palatino Linotype', bold=True, color="FFFFFF", size=11)
        self.body_font = Font(name='Palatino Linotype', size=10)
        self.subtotal_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.tender_metrics = {}

    def find_column(self, columns: List[str], col_type: str) -> Optional[str]:
        """Find column by type with case-insensitive matching"""
        possible_names = self.column_mapping[col_type]

        for col in columns:
            col_lower = str(col).lower().strip()
            for name in possible_names:
                if name in col_lower:
                    return col

        return None

    def _clean_numeric_series(self, series: pd.Series) -> pd.Series:
        """
        Clean numeric-like strings: remove commas, non-breaking spaces, then convert to numeric.
        Returns float series with NaN for unparseable values.
        """
        s = series.astype(str).str.replace(',', '', regex=False).str.replace(u'\xa0', '', regex=False).str.strip()
        s = s.replace({'': np.nan, 'nan': np.nan, 'None': np.nan})
        return pd.to_numeric(s, errors='coerce')

    def read_tender_file(self, file_path: str, tender_name: str) -> Optional[pd.DataFrame]:
        """Read and process a tender file"""
        try:
            read_kwargs = dict(
                skiprows=5,
                encoding='utf-8',
                na_values=['', 'NA', 'NaN', 'NULL', 'nan', 'null'],
                keep_default_na=True,
                low_memory=False,
                engine='c'
            )
            try:
                df = pd.read_csv(file_path, **read_kwargs)
            except Exception:
                read_kwargs['encoding'] = 'ISO-8859-1'
                df = pd.read_csv(file_path, **read_kwargs)

            if df.empty:
                return None

            df.columns = df.columns.str.strip()

            store_id_col = self.find_column(df.columns.tolist(), 'store_id')
            store_response_col = self.find_column(df.columns.tolist(), 'store_response')
            auto_approved_col = self.find_column(df.columns.tolist(), 'auto_approved')

            if not all([store_id_col, store_response_col, auto_approved_col]):
                return None

            keep_cols = [store_id_col, store_response_col, auto_approved_col]

            sales_date_col = None
            for col in df.columns:
                col_lower = str(col).lower().strip()
                if 'sales' in col_lower and 'date' in col_lower:
                    sales_date_col = col
                    keep_cols.append(col)
                    break

            df = df[keep_cols].copy()

            rename_dict = {
                store_id_col: 'Store_ID',
                store_response_col: 'Store_Response_Entry',
                auto_approved_col: 'Auto_Approved_Date'
            }
            if sales_date_col:
                rename_dict[sales_date_col] = 'Sales_Date'

            df = df.rename(columns=rename_dict)

            # Clean numeric store id and response columns robustly
            df['Store_ID'] = pd.to_numeric(df['Store_ID'].astype(str).str.replace(',', '', regex=False).str.strip(), errors='coerce')
            df['Store_Response_Entry'] = self._clean_numeric_series(df['Store_Response_Entry'])

            df = df.dropna(subset=['Store_ID', 'Store_Response_Entry'], how='any')
            df = df[df['Store_Response_Entry'] != 0]

            df['Auto_Approved_Date'] = df['Auto_Approved_Date'].astype(str).str.strip()
            df = df[~df['Auto_Approved_Date'].isin(['nan', 'NaN', 'NA', 'NULL', '', '0', 'None'])]
            df = df[df['Auto_Approved_Date'].str.len() > 0]

            df['Store_ID'] = df['Store_ID'].astype('int32')
            df['Tender_Type'] = tender_name

            if df.empty:
                return None

            gc.collect()
            return df

        except Exception:
            return None

    def find_and_remove_netting_items(self, store_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Find items that net off and remove them.

        Enhancements:
        - Pair-wise netting (existing)
        - Multi-entry netting within same Sales_Date and same Tender_Type -> 'Within-Tender-Multiple entries'
        - Multi-entry cross-tender netting within same Sales_Date -> 'Cross-Tender-Multiple entries'
        """
        if len(store_df) <= 1:
            total = float(abs(store_df['Store_Response_Entry'].sum())) if len(store_df) == 1 else 0.0
            if total >= 100:
                return store_df.copy(), pd.DataFrame()
            else:
                return pd.DataFrame(), pd.DataFrame()

        df = store_df.reset_index(drop=True).copy()
        df['abs_value'] = df['Store_Response_Entry'].abs()
        df = df.sort_values('abs_value', ascending=False).reset_index(drop=True)

        used_flags = [False] * len(df)
        netting_records = []

        # Helper to mark indices used
        def mark_used(indices: List[int]):
            for idx in indices:
                used_flags[idx] = True

        # 1) Existing pair-wise greedy netting (largest with next entries)
        for i in range(len(df)):
            if used_flags[i]:
                continue
            # Use iloc for positional indexing to avoid issues
            val_i = float(df.iloc[i]['Store_Response_Entry'])
            for j in range(i + 1, len(df)):
                if used_flags[j]:
                    continue
                val_j = float(df.iloc[j]['Store_Response_Entry'])
                combined_variance = abs(val_i + val_j)
                if combined_variance < self.netting_off_threshold:
                    # Mark used and record pair
                    mark_used([i, j])
                    netting_records.append({
                        'Store_ID': int(df.iloc[i]['Store_ID']),
                        'Sales_Date': str(df.iloc[i].get('Sales_Date', '')),
                        'Tender_Type_1': str(df.iloc[i]['Tender_Type']),
                        'Store_Response_Entry_1': val_i,
                        'Tender_Type_2': str(df.iloc[j]['Tender_Type']),
                        'Store_Response_Entry_2': val_j,
                        'Combined_Variance': combined_variance,
                        'Netting_Type': 'Cross-Tender' if df.iloc[i]['Tender_Type'] != df.iloc[j]['Tender_Type'] else 'Within-Tender',
                        'Netting_Members': f"{i},{j}"
                    })
                    break

        # 2) Multi-entry netting by Sales_Date and Tender (Within-Tender-Multiple entries)
        if 'Sales_Date' in df.columns:
            for (sales_date, tender_type), grp in df[~pd.Series(used_flags)].groupby(['Sales_Date', 'Tender_Type']):
                if len(grp) < 2:
                    continue
                idxs = grp.index.tolist()
                values = grp['Store_Response_Entry'].astype(float)
                group_sum = float(values.sum())
                if abs(group_sum) < self.netting_off_threshold:
                    # mark and record group netting
                    mark_used(idxs)
                    netting_records.append({
                        'Store_ID': int(df.iloc[idxs[0]]['Store_ID']),
                        'Sales_Date': str(sales_date),
                        'Tender_Type_1': str(tender_type),
                        'Store_Response_Entry_1': None,
                        'Tender_Type_2': None,
                        'Store_Response_Entry_2': None,
                        'Combined_Variance': abs(group_sum),
                        'Netting_Type': 'Within-Tender-Multiple entries',
                        'Netting_Members': ",".join(map(str, idxs)),
                        'Group_Sum': group_sum
                    })

        # 3) Multi-entry cross-tender netting on same Sales_Date (Cross-Tender-Multiple entries)
        if 'Sales_Date' in df.columns:
            for sales_date, grp in df[~pd.Series(used_flags)].groupby(['Sales_Date']):
                if len(grp) < 2:
                    continue
                idxs = grp.index.tolist()
                values = grp['Store_Response_Entry'].astype(float)
                group_sum = float(values.sum())
                if abs(group_sum) < self.netting_off_threshold:
                    mark_used(idxs)
                    # Identify if mixed tenders -> cross-tender
                    tender_types = grp['Tender_Type'].unique().tolist()
                    netting_records.append({
                        'Store_ID': int(df.iloc[idxs[0]]['Store_ID']),
                        'Sales_Date': str(sales_date),
                        'Tender_Type_1': ",".join(str(t) for t in tender_types),
                        'Store_Response_Entry_1': None,
                        'Tender_Type_2': None,
                        'Store_Response_Entry_2': None,
                        'Combined_Variance': abs(group_sum),
                        'Netting_Type': 'Cross-Tender-Multiple entries' if len(tender_types) > 1 else 'Within-Tender-Multiple entries',
                        'Netting_Members': ",".join(map(str, idxs)),
                        'Group_Sum': group_sum
                    })

        # Build netting reference DataFrame
        netting_ref_df = pd.DataFrame(netting_records) if netting_records else pd.DataFrame()

        # Remaining exceptional items = those not used
        remaining_idxs = [i for i, used in enumerate(used_flags) if not used]
        if remaining_idxs:
            exceptional_df = df.loc[remaining_idxs].copy()
            if 'abs_value' in exceptional_df.columns:
                exceptional_df = exceptional_df.drop(columns=['abs_value'])
            # final check: ensure store-level remaining sum >= 100 to be deemed exception
            total_remaining = float(abs(exceptional_df['Store_Response_Entry'].sum()))
            if total_remaining >= 100 - 1e-6:
                return exceptional_df.reset_index(drop=True), netting_ref_df

        return pd.DataFrame(), netting_ref_df

    def process_all_tenders(self, tender_files: Dict[str, str]) -> Optional[Dict]:
        """Process all tender files with netting logic"""
        all_data = {}
        processed_tenders = []

        for tender_name, file_path in tender_files.items():
            df = self.read_tender_file(file_path, tender_name)

            if df is not None and not df.empty:
                all_data[tender_name] = df
                processed_tenders.append(tender_name)

                self.tender_metrics[tender_name] = {
                    'total_entries': len(df),
                    'exception_entries': 0,
                    'cross_tender_netting': 0,
                    'within_tender_netting': 0,
                    'total_netting_variance': 0.0
                }

        if not all_data:
            return None

        combined_data = pd.concat([all_data[t] for t in processed_tenders],
                                  ignore_index=True,
                                  copy=False)

        # Ensure Store_Response_Entry numeric (strip commas etc.)
        combined_data['Store_Response_Entry'] = self._clean_numeric_series(combined_data['Store_Response_Entry']).fillna(0.0)

        # Group totals by store
        store_totals = combined_data.groupby('Store_ID', as_index=True, observed=True)['Store_Response_Entry'].sum()

        # Store-level exceptions: total absolute >= 100 only
        exception_mask = (abs(store_totals) >= 100)
        exception_stores = store_totals[exception_mask]

        if len(exception_stores) == 0:
            return {
                'summary': pd.DataFrame(),
                'classification': pd.DataFrame(),
                'exceptions': {t: pd.DataFrame() for t in processed_tenders},
                'netting_reference': pd.DataFrame(),
                'tender_performance': pd.DataFrame(),
                'tender_names': processed_tenders,
                'total_stores': len(store_totals),
                'exception_stores': 0
            }

        exception_store_ids = exception_stores.index.tolist()
        filtered_exception_data = {}
        all_netting_reference = []

        for store_id in exception_store_ids:
            store_all_tenders = combined_data[combined_data['Store_ID'] == store_id].copy()
            # Ensure numeric safety
            store_all_tenders['Store_Response_Entry'] = pd.to_numeric(store_all_tenders['Store_Response_Entry'], errors='coerce').fillna(0.0)

            exceptional_items, netting_ref = self.find_and_remove_netting_items(store_all_tenders)

            if netting_ref is not None and not netting_ref.empty:
                all_netting_reference.append(netting_ref)
                for idx, row in netting_ref.iterrows():
                    # Update tender_metrics aggregates where possible
                    tender_type_str = row.get('Tender_Type_1', None)
                    netting_type = row.get('Netting_Type', '')
                    variance = float(row.get('Combined_Variance', 0.0) or 0.0)
                    # If tender_type_str is a comma-separated list, increment metrics for each tender if exists
                    if tender_type_str:
                        for tname in str(tender_type_str).split(','):
                            tname = tname.strip()
                            if tname in self.tender_metrics:
                                if 'Cross-Tender' in netting_type or 'Cross-Tender-Multiple' in netting_type:
                                    self.tender_metrics[tname]['cross_tender_netting'] += 1
                                else:
                                    self.tender_metrics[tname]['within_tender_netting'] += 1
                                self.tender_metrics[tname]['total_netting_variance'] += variance

            if exceptional_items is not None and not exceptional_items.empty:
                total_exc = float(abs(exceptional_items['Store_Response_Entry'].sum()))
                if total_exc >= 100 - 1e-6:
                    for tender_name in processed_tenders:
                        tender_items = exceptional_items[exceptional_items['Tender_Type'] == tender_name]
                        if not tender_items.empty:
                            self.tender_metrics[tender_name]['exception_entries'] += len(tender_items)
                            if tender_name not in filtered_exception_data:
                                filtered_exception_data[tender_name] = []
                            filtered_exception_data[tender_name].append(tender_items)

        for tender_name in processed_tenders:
            if tender_name in filtered_exception_data and filtered_exception_data[tender_name]:
                filtered_exception_data[tender_name] = pd.concat(
                    filtered_exception_data[tender_name],
                    ignore_index=True
                )
            else:
                filtered_exception_data[tender_name] = pd.DataFrame()

        netting_reference_df = pd.concat(all_netting_reference, ignore_index=True) if all_netting_reference else pd.DataFrame()

        total_auto_updated_responses = combined_data.groupby('Store_ID', as_index=True).size()

        summary_data = []
        for store_id in exception_store_ids:
            store_all_tenders = combined_data[combined_data['Store_ID'] == store_id].copy()
            exceptional_items, _ = self.find_and_remove_netting_items(store_all_tenders)

            if exceptional_items is not None and not exceptional_items.empty:
                new_total = float(exceptional_items['Store_Response_Entry'].sum())
                if abs(new_total) >= 100 - 1e-6:
                    line_item_count = len(exceptional_items)
                    summary_row = {
                        'Store_ID': store_id,
                        'Total_Auto-Updated_Responses': int(total_auto_updated_responses.get(store_id, 0)),
                        'Exceptional_Line_Items': line_item_count,
                        'Sum_of_Exceptional_Responses': new_total,
                        'Classification': self.classify_total_response(new_total),
                        'Error_Rate_%': (line_item_count / total_auto_updated_responses.get(store_id, 1) * 100)
                                        if total_auto_updated_responses.get(store_id, 0) > 0 else 0
                    }

                    for tender_name in processed_tenders:
                        tender_items = exceptional_items[exceptional_items['Tender_Type'] == tender_name]
                        tender_sum = float(tender_items['Store_Response_Entry'].sum()) if not tender_items.empty else 0.0
                        summary_row[f'{tender_name}_Sum'] = tender_sum

                    summary_data.append(summary_row)

        summary_df = pd.DataFrame(summary_data)

        if not summary_df.empty:
            classification_counts = summary_df['Classification'].value_counts().reset_index()
            classification_counts.columns = ['Classification', 'Store_Count']

            classification_order = [
                "Diff less than +/- 100",
                "Diff b/w +/- 1000",
                "Diff b/w +/- 5000",
                "Diff b/w +/- 10000",
                "Diff b/w +/- 25000",
                "Diff b/w +/- 50000",
                "Diff more than +/- 50000"
            ]

            classification_counts['Classification'] = pd.Categorical(
                classification_counts['Classification'],
                categories=classification_order,
                ordered=True
            )
            classification_counts = classification_counts.sort_values('Classification').reset_index(drop=True)
        else:
            classification_counts = pd.DataFrame()

        tender_performance_data = []
        for tender_name in processed_tenders:
            metrics = self.tender_metrics.get(tender_name, {})
            tender_performance_data.append({
                'Tender': tender_name,
                'Total_Entries': metrics.get('total_entries', 0),
                'Exceptional_Entries': metrics.get('exception_entries', 0),
                'Exception_Rate_%': (metrics.get('exception_entries', 0) / metrics.get('total_entries', 1) * 100) if metrics.get('total_entries', 0) > 0 else 0,
                'Items_Removed_by_Netting': metrics.get('cross_tender_netting', 0) + metrics.get('within_tender_netting', 0),
                'Total_Netting_Variance': metrics.get('total_netting_variance', 0.0)
            })

        tender_performance_df = pd.DataFrame(tender_performance_data)

        gc.collect()

        return {
            'summary': summary_df,
            'classification': classification_counts,
            'exceptions': filtered_exception_data,
            'netting_reference': netting_reference_df,
            'tender_performance': tender_performance_df,
            'tender_names': processed_tenders,
            'total_stores': len(store_totals),
            'exception_stores': len(exception_stores)
        }

    def classify_total_response(self, total: float) -> str:
        """Classify total based on thresholds"""
        abs_total = abs(total)

        for min_val, max_val, label in self.classification_thresholds:
            if min_val <= abs_total <= max_val:
                return label

        return "Diff more than +/- 50000"

    def format_worksheet(self, worksheet, has_data: bool = True, start_row: int = 3):
        """Format worksheet with professional styling"""
        worksheet.sheet_view.showGridLines = False

        if not has_data:
            return

        # Format header row (at start_row)
        for cell in worksheet[start_row]:
            if cell.value:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.border

        # Format data rows
        for row in worksheet.iter_rows(min_row=start_row + 1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.font = self.body_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center')

                try:
                    if cell.value and isinstance(cell.value, (int, float)):
                        col_header = worksheet.cell(row=start_row, column=cell.column).value

                        if col_header and ('Error_Rate_%' in str(col_header) or 'Exception_Rate_%' in str(col_header)):
                            cell.number_format = '0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        elif col_header and any(x in str(col_header) for x in ['Response', 'Sum', 'Items', 'Variance', 'Entries']):
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                except:
                    pass

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            try:
                column_letter = get_column_letter(column[0].column)
            except:
                continue

            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    max_length = max(max_length, len(cell_value))
                except:
                    pass

            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def add_subtotal_row(self, worksheet, header_row: int, data_start_row: int, data_end_row: int):
        """
        Write SUBTOTAL function row at row 2 (above header) WITHOUT inserting rows,
        so header remains at header_row (e.g. 3).
        """
        subtotal_row = header_row - 1  # row above header (usually 2)

        # Identify numeric columns from header
        numeric_cols = []
        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value:
                col_letter = get_column_letter(col_idx)
                col_header = str(cell.value).lower()
                if any(x in col_header for x in ['response', 'sum', 'variance', 'items', 'entries']):
                    numeric_cols.append((col_letter, col_idx))

        # Add SUBTOTAL function to numeric columns, writing directly into subtotal_row
        for col_letter, col_idx in numeric_cols:
            cell = worksheet[f'{col_letter}{subtotal_row}']
            cell.value = f'=SUBTOTAL(9,{col_letter}{data_start_row}:{col_letter}{data_end_row})'
            cell.font = Font(name='Palatino Linotype', bold=True, size=10)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = self.border
            cell.fill = self.subtotal_fill

        # Add label in first column of subtotal row
        first_cell = worksheet[f'A{subtotal_row}']
        first_cell.value = 'SUBTOTAL'
        first_cell.font = Font(name='Palatino Linotype', bold=True, size=10)
        first_cell.border = self.border
        first_cell.fill = self.subtotal_fill

    def save_to_excel(self, results: Dict, output_path: str):
        """
        Save results to Excel.
        'User_Manual' and 'Tender_Performance' sheets are intentionally omitted.
        Header row will be at row 3 and SUBTOTAL formulas will be placed at row 2.
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Store Summary
                if not results['summary'].empty:
                    results['summary'].to_excel(writer, sheet_name='Store_Summary', index=False, startrow=2)
                    store_summary_ws = writer.sheets['Store_Summary']
                    header_row = 3
                    data_start = header_row + 1
                    data_end = data_start + len(results['summary']) - 1 if len(results['summary']) > 0 else data_start
                    # place subtotals in row 2 (no insertion)
                    self.add_subtotal_row(store_summary_ws, header_row, data_start, data_end)
                    self.format_worksheet(store_summary_ws, has_data=True, start_row=header_row)
                else:
                    pd.DataFrame({"Message": ["No real exceptions found"]}).to_excel(
                        writer, sheet_name='Store_Summary', index=False, startrow=2)

                # Classification Summary
                if not results['classification'].empty:
                    results['classification'].to_excel(writer, sheet_name='Classification_Summary', index=False, startrow=2)
                    class_summary_ws = writer.sheets['Classification_Summary']
                    header_row = 3
                    data_start = header_row + 1
                    data_end = data_start + len(results['classification']) - 1 if len(results['classification']) > 0 else data_start
                    self.add_subtotal_row(class_summary_ws, header_row, data_start, data_end)
                    self.format_worksheet(class_summary_ws, has_data=True, start_row=header_row)
                else:
                    pd.DataFrame({"Message": ["No real exceptions found"]}).to_excel(
                        writer, sheet_name='Classification_Summary', index=False, startrow=2)

                # Netting Reference
                if not results['netting_reference'].empty:
                    results['netting_reference'].to_excel(writer, sheet_name='Netting_Reference', index=False, startrow=2)
                    netting_ref_ws = writer.sheets['Netting_Reference']
                    header_row = 3
                    data_start = header_row + 1
                    data_end = data_start + len(results['netting_reference']) - 1 if len(results['netting_reference']) > 0 else data_start
                    self.add_subtotal_row(netting_ref_ws, header_row, data_start, data_end)
                    self.format_worksheet(netting_ref_ws, has_data=True, start_row=header_row)
                else:
                    pd.DataFrame({"Message": ["No netting detected"]}).to_excel(
                        writer, sheet_name='Netting_Reference', index=False, startrow=2)

                # Exceptions per tender
                for tender_name, df in results['exceptions'].items():
                    if df is not None and not df.empty:
                        sheet_name = f"{tender_name}_Exceptions"
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:28] + "..."
                        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                        exception_ws = writer.sheets[sheet_name]
                        header_row = 3
                        data_start = header_row + 1
                        data_end = data_start + len(df) - 1 if len(df) > 0 else data_start
                        self.add_subtotal_row(exception_ws, header_row, data_start, data_end)
                        self.format_worksheet(exception_ws, has_data=True, start_row=header_row)

            return True

        except Exception:
            return False

    def add_tender_performance_chart(self, worksheet, tender_data):
        """Add performance chart (kept for completeness; not used when sheet omitted)"""
        try:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Tender Performance Summary"
            chart.y_axis.title = "Count / Percentage"
            chart.x_axis.title = "Tender Type"

            data = Reference(worksheet, min_col=2, min_row=3, max_row=len(tender_data) + 3, max_col=3)
            categories = Reference(worksheet, min_col=1, min_row=4, max_row=len(tender_data) + 3)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 12
            chart.width = 20

            worksheet.add_chart(chart, "G4")
        except:
            pass
